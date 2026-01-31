#!/usr/bin/env python3
"""
Scrape all episode titles + air dates from TVDB "All seasons / Official" page
and write an Excel table with the requested columns.

Output columns:
- SeasonEpisode: SxxEyy where xx is the season number starting at 01,
                 and yy is the episode number within that season.
- Broadcast Date: yyyy-mm-dd
- Absolute episode number (sorted by season/episode)
- Title as mentioned in TVDB
- Title in the form: "SxxEyy - Mit dem Zug durch ... .mp4" ( ... = Title )

Requires:
  pip install requests beautifulsoup4 openpyxl
"""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from datetime import datetime
from typing import List, Optional, Tuple

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry


DEFAULT_URL = "https://www.thetvdb.com/series/280509-show/allseasons/official"
DEFAULT_OUT_XLSX = "tvdb_episodes_mit_dem_zug_durch.xlsx"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/120.0 Safari/537.36"
    )
}

SXXEYY_RE = re.compile(r"\bS(\d{1,3})E(\d{1,4})\b", re.IGNORECASE)

# Windows-forbidden filename chars: \ / : * ? " < > |
BAD_FILENAME_CHARS_RE = re.compile(r'[\\/:*?"<>|]')


@dataclass(frozen=True)
class Episode:
    season: int
    episode: int
    airdate_iso: Optional[str]  # yyyy-mm-dd or None
    title: str


def clean_spaces(s: str) -> str:
    return re.sub(r"\s+", " ", s).strip()


def sanitize_filename_component(s: str) -> str:
    s = clean_spaces(s)
    s = BAD_FILENAME_CHARS_RE.sub(" ", s)
    s = clean_spaces(s)

    # Windows also dislikes trailing dots/spaces. Keep it simple.
    s = s.rstrip(" .")

    # Optional: cap length to avoid path issues (safe default)
    if len(s) > 140:
        s = s[:140].rstrip(" .")

    return s


def parse_airdate_to_iso(date_text: str) -> Optional[str]:
    """
    TVDB often shows: 'July 10, 2006'
    Sometimes:       'Jul 10, 2006'
    Convert to '2006-07-10'. If parsing fails, return None.
    """
    date_text = clean_spaces(date_text)

    # Common non-date placeholders
    lowered = date_text.lower()
    if lowered in {"tba", "n/a", "na", "-", "—", ""}:
        return None

    for fmt in ("%B %d, %Y", "%b %d, %Y"):
        try:
            dt = datetime.strptime(date_text, fmt)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            continue

    return None


def make_session() -> requests.Session:
    """
    Create a Session with retries/backoff for transient errors and rate limits.
    """
    s = requests.Session()

    retry = Retry(
        total=5,
        connect=5,
        read=5,
        backoff_factor=0.6,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=frozenset(["GET"]),
        raise_on_status=False,
    )
    adapter = HTTPAdapter(max_retries=retry)
    s.mount("https://", adapter)
    s.mount("http://", adapter)

    s.headers.update(HEADERS)
    return s


def fetch_html(url: str, timeout_s: int = 30) -> str:
    with make_session() as s:
        r = s.get(url, timeout=timeout_s)
        r.raise_for_status()
        return r.text


def _extract_airdate_from_container(container) -> Optional[str]:
    """
    More robust airdate extraction: scan text fragments within the container
    and return the first parsable date.
    """
    # Get a list of text chunks (avoid concatenating everything into one line)
    texts: List[str] = []
    for t in container.stripped_strings:
        chunk = clean_spaces(t)
        if chunk:
            texts.append(chunk)

    # Try each chunk directly
    for chunk in texts:
        iso = parse_airdate_to_iso(chunk)
        if iso:
            return iso

    # Heuristic: sometimes date is embedded in a longer string, so try regex pull
    # for patterns like "July 10, 2006" or "Jul 10, 2006".
    date_like_re = re.compile(
        r"\b([A-Za-z]{3,9})\s+(\d{1,2}),\s+(\d{4})\b"
    )
    for chunk in texts:
        m = date_like_re.search(chunk)
        if m:
            candidate = f"{m.group(1)} {m.group(2)}, {m.group(3)}"
            iso = parse_airdate_to_iso(candidate)
            if iso:
                return iso

    return None


def extract_episodes(html: str, debug: bool = False) -> List[Episode]:
    soup = BeautifulSoup(html, "html.parser")
    episodes: List[Episode] = []

    # Still based on h4 markers, but improved airdate scanning + debug diagnostics.
    h4s = soup.find_all("h4")
    if debug:
        print(f"[debug] Found {len(h4s)} <h4> elements total.", file=sys.stderr)

    for h4 in h4s:
        h4_text = clean_spaces(h4.get_text(" ", strip=True))
        m = SXXEYY_RE.search(h4_text)
        if not m:
            continue

        season = int(m.group(1))
        epnum = int(m.group(2))

        # Title: prefer <a> inside the h4
        a = h4.find("a")
        title = clean_spaces(a.get_text(" ", strip=True)) if a else ""
        if not title:
            # Fallback: remove SxxEyy from header text
            title = clean_spaces(SXXEYY_RE.sub("", h4_text)).strip("-–—|• ").strip()

        # Air date: scan the closest meaningful container
        airdate_iso: Optional[str] = None
        container_li = h4.find_parent("li")
        if container_li:
            airdate_iso = _extract_airdate_from_container(container_li)

        if debug and not airdate_iso:
            # Print a small snippet for troubleshooting (not the full page).
            snippet = clean_spaces(container_li.get_text(" ", strip=True))[:200] if container_li else ""
            print(
                f"[debug] No airdate parsed for S{season:02d}E{epnum:02d}. "
                f"Container snippet: {snippet!r}",
                file=sys.stderr,
            )

        episodes.append(Episode(season=season, episode=epnum, airdate_iso=airdate_iso, title=title))

    episodes.sort(key=lambda e: (e.season, e.episode))
    return episodes


def write_excel(episodes: List[Episode], out_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Episodes"

    headers = [
        "SeasonEpisode",
        "Broadcast Date",
        "Absolute episode number (sorted S/E)",
        "Title (TVDB)",
        'Title as "SxxEyy - Mit dem Zug durch ... .mp4"',
    ]
    ws.append(headers)

    for idx, e in enumerate(episodes, start=1):
        season_episode = f"S{e.season:02d}E{e.episode:02d}"
        safe_title = sanitize_filename_component(e.title)
        filename_title = f"{season_episode} - Mit dem Zug durch {safe_title}.mp4"

        ws.append([
            season_episode,
            e.airdate_iso or "",
            idx,
            e.title,
            filename_title,
        ])

    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Column widths
    col_widths = [16, 16, 30, 45, 60]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    # Add an Excel "Table" (filterable)
    last_row = ws.max_row
    last_col = ws.max_column
    table_ref = f"A1:{get_column_letter(last_col)}{last_row}"
    table = Table(displayName="EpisodesTable", ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    wb.save(out_path)


def print_summary(episodes: List[Episode]) -> None:
    missing = sum(1 for e in episodes if not e.airdate_iso)
    seasons = {}
    for e in episodes:
        seasons[e.season] = seasons.get(e.season, 0) + 1

    first_date = min((e.airdate_iso for e in episodes if e.airdate_iso), default=None)
    last_date = max((e.airdate_iso for e in episodes if e.airdate_iso), default=None)

    season_parts = ", ".join(f"S{s:02d}:{n}" for s, n in sorted(seasons.items()))
    print(f"Episodes parsed: {len(episodes)}")
    print(f"Missing air dates: {missing}")
    if first_date and last_date:
        print(f"Air date range: {first_date} → {last_date}")
    print(f"Count by season: {season_parts}")


def parse_args(argv: Optional[List[str]] = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Scrape TVDB episodes and write an Excel file.")
    p.add_argument("--url", default=DEFAULT_URL, help="TVDB 'all seasons / official' URL")
    p.add_argument("-o", "--out", default=DEFAULT_OUT_XLSX, help="Output .xlsx path")
    p.add_argument("--timeout", type=int, default=30, help="HTTP timeout in seconds (default: 30)")
    p.add_argument("--debug", action="store_true", help="Print debug diagnostics to stderr")
    return p.parse_args(argv)


def main(argv: Optional[List[str]] = None) -> int:
    args = parse_args(argv)

    try:
        html = fetch_html(args.url, timeout_s=args.timeout)
    except requests.RequestException as e:
        print(f"ERROR: Failed to fetch page: {e}", file=sys.stderr)
        return 2

    episodes = extract_episodes(html, debug=args.debug)

    if not episodes:
        print(
            "No episodes parsed. TVDB may have changed markup or blocked the request.\n"
            "Tips:\n"
            "- Try again later (rate limiting happens).\n"
            "- Try a different URL variant.\n"
            "- If the site requires login/cookies, you may need an authenticated session.\n"
            "- Run with --debug to inspect parsing hints.\n",
            file=sys.stderr,
        )
        return 1

    write_excel(episodes, args.out)
    print_summary(episodes)
    print(f"Wrote {len(episodes)} rows to {args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
